from openpyxl import load_workbook


class moviesoperation:
    def Movies(self):
        id = 0
        print("--- Movies List ---")
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        rows = len(list(ws.rows))
        range = ws["A1":"A" + str(rows)]
        for movie in range:
            for x in movie:
                id += 1
                print(id, " ", x.value)

    def MovieInfo(self, choice):
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        range = ws[str(choice)]
        print("-----Movie Details-----")
        print("-Title-    -Genre-    -length-     -Cast-      -Capacity-     -AdminRatings-")
        for movies in range:
            print(movies.value, end="     ")
        wb.save("movies.xlsx")
        print()

    def CancelTickets(self, choice):
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        cancel = int(input("Enter Number of seats you want to cancel: "))
        ws["E" + str(choice)] = ws["E" + str(choice)].value + cancel
        wb.save("movies.xlsx")

    def UserRatings(self, choice):
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        user_rating = input("Enter ratings for the movie: ")
        ws["G" + str(choice)] = user_rating
        wb.save("movies.xlsx")

    def BookTickets(self, choice):
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        print("Select Timings For Your Show: ")
        print("1 -  8:00-10:00")
        print("2 -  10:30-12:30")
        print("3 -  1:00-3:00")
        print("4 -  3:30-5:30")
        print("5 -  6:00-8:30")
        booktime = input()
        print("Seats Remaining: ", int(ws["E" + str(choice)].value))
        bookseat = int(input("Enter Number of seats you want to book: "))
        ws["E" + str(choice)] = int(ws["E" + str(choice)].value) - bookseat
        print("Tickets Booked Successfully")
        wb.save("movies.xlsx")