from openpyxl import load_workbook


class moviescrud:
    def AddMovie(self):
        global Title
        print("-----Add Movie-----")
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        rows = len(list(ws.rows))
        try:
            number = int(input("How many movies do you want to add?: "))
        except ValueError as e:
            print("Please enter valid detail!!")
            return False
        for i in range(rows, number + rows):
            Title = input("Enter Title of the movie: ")
            Genre = input("Enter Genre of the movie: ")
            Length = input("Enter Length of the movie: ")
            Cast = input("Enter Cast of the movie: ")
            Capacity = input("Enter Capacity of the movie: ")
            AdminRatings = input("Enter Admin Ratings out of 10 for the movie: ")
            ws.append([Title, Genre, Length, Cast, Capacity, AdminRatings])
        print(Title + " Movie Added Successfully")
        wb.save("movies.xlsx")
        print()
        return True


    def EditMovie(self):
        global Title
        print("-----Edit Movie-----")
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        rows = len(list(ws.rows))
        edit = input("Enter the title of the movie you want to edit: ")
        range = ws["A1":"A" + str(rows)]
        row_number = 0
        for movie in range:
            for x in movie:
                row_number += 1
                if x.value == edit:
                    Title = input("Enter New title: ")
                    ws["A" + str(row_number)] = Title
                    Genre = input("enter new genre: ")
                    ws["B" + str(row_number)] = Genre
                    Length = input("Enter new length: ")
                    ws["C" + str(row_number)] = Length
                    Cast = input("Enter new cast: ")
                    ws["D" + str(row_number)] = Cast
                    Capacity = input("Enter new capacity: ")
                    ws["E" + str(row_number)] = Capacity
                    Adminrating = input("Enter new admin rating: ")
                    ws["F" + str(row_number)] = Adminrating
                    break
        print(Title + " Movie Details updated successfully")
        wb.save("movies.xlsx")
        print()

    def DeleteMovie(self):
        print("-----Delete Movie-----")
        wb = load_workbook("movies.xlsx")
        ws = wb.active
        rows = len(list(ws.rows))
        delete = input("Enter the title of the movie you want to delete: ")
        range = ws["A1":"A" + str(rows)]
        row_number = 0
        for movie in range:
            for x in movie:
                row_number += 1
                if x.value == delete:
                    ws.delete_rows(row_number)
                    break
        print(Title + " Movie deleted successfully")
        wb.save("movies.xlsx")
        print()